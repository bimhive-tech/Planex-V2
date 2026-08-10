"""Tests for the Planex Code P6 parser (p6_id_schedule_import.py).

Built against the team's first real export (Mansoura 6 - Building, Aug
2026) — see that module's docstring for the empirically-confirmed shape.
"""
import datetime
import io

import openpyxl
from django.test import TestCase

from apps.accounts.models import Company

from .imports import import_workbook
from .models import Activity, Project, ProjectScope
from .p6_id_schedule_import import is_milestone_code, milestone_scope_path, segment_path


class SegmentPathTests(TestCase):
    def test_real_file_shape_ten_segments_two_dropped(self):
        """The team's actual export: 10 dash-separated segments (not the
        legend's full 12 — Level and Sub-discipline are dropped entirely),
        with Area/Sub-area/Part kept as a literal "0" placeholder."""
        path = segment_path("MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1")
        self.assertEqual(path, ["PH1", "Z(A)", "Building 6", "Internal Finishes"])

    def test_bare_tag_word_is_a_placeholder(self):
        # "CON" alone (no distinguishing suffix) means "this project doesn't
        # branch on Construction" — same as "0" for Area/Sub-area/Part.
        path = segment_path("MN(6)-CON-0-0-PH2-Z(B)-0-Building 12-Stairs-30")
        self.assertEqual(path, ["PH2", "Z(B)", "Building 12", "Stairs"])

    def test_project_code_and_differentiator_are_dropped(self):
        # First segment (project code) and last (per-row differentiator) are
        # never part of the tree path — the real Activity ID is used as the
        # leaf's own code instead of the differentiator.
        path = segment_path("PN01-PH1-1")
        self.assertEqual(path, ["PH1"])

    def test_all_placeholders_returns_empty(self):
        self.assertEqual(segment_path("MN(6)-CON-0-0-0-0-0-0-DEC-1"), [])

    def test_too_few_segments_returns_empty(self):
        self.assertEqual(segment_path("MN(6)-1"), [])
        self.assertEqual(segment_path("not a code"), [])

    def test_non_string_returns_empty(self):
        self.assertEqual(segment_path(None), [])
        self.assertEqual(segment_path(123), [])


class MilestoneCodeTests(TestCase):
    def test_bare_project_wide_milestone_code(self):
        self.assertTrue(is_milestone_code("MN(6)-MS-1"))
        self.assertEqual(milestone_scope_path("MN(6)-MS-1"), ())

    def test_building_tied_milestone_code(self):
        code = "MN(6)-MS-PH1-Z(A)-Building 15-1"
        self.assertTrue(is_milestone_code(code))
        self.assertEqual(milestone_scope_path(code), ("PH1", "Z(A)", "Building 15"))

    def test_discipline_code_is_not_a_milestone_code(self):
        self.assertFalse(is_milestone_code("MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1"))

    def test_non_string_is_not_a_milestone_code(self):
        self.assertFalse(is_milestone_code(None))
        self.assertFalse(is_milestone_code(123))


class IdScheduleImportTests(TestCase):
    """Same header shape as the reference P6 schedule export, plus a
    separate "Planex Code" column — Activity ID/Name stay the leaf
    activity's own identity; Planex Code only drives the WBS tree."""

    HEADER = ["Planex Code", "Activity ID", "Activity Name", "Original Duration", "Start", "Finish",
              "Activity % Complete", "Budgeted Total Cost", "Earned Value Cost"]

    def _workbook(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(self.HEADER)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_builds_tree_from_real_file_shaped_codes(self):
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1", "MN6-A6-01-01", "Seal",
             10, d(2026, 1, 1), d(2026, 1, 10), 0.5, 1000, 500],
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-2", "MN6-A6-01-02", "Putty",
             5, d(2026, 1, 11), d(2026, 1, 15), 0, 500, 0],
            # A second discipline under the same Building.
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Stairs-1", "MN6-A6-02-01", "Stair render",
             8, d(2026, 2, 1), d(2026, 2, 8), 1, 2000, 2000],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")

        result = import_workbook(project, self._workbook(rows), source="planex_code.xlsx")
        self.assertEqual(result["source_kind"], "p6_schedule")
        self.assertEqual(result["activities"], 3)

        stage = ProjectScope.objects.get(project=project, name="PH1")
        self.assertEqual(stage.scope_type, ProjectScope.ScopeType.STAGE)
        self.assertEqual(stage.parent, None)

        zone = ProjectScope.objects.get(project=project, name="Z(A)")
        self.assertEqual(zone.scope_type, ProjectScope.ScopeType.ZONE)
        self.assertEqual(zone.parent, stage)

        building = ProjectScope.objects.get(project=project, name="Building 6")
        self.assertEqual(building.scope_type, ProjectScope.ScopeType.AREA)
        self.assertEqual(building.parent, zone)

        disciplines = list(ProjectScope.objects.filter(project=project, parent=building).order_by("name"))
        self.assertEqual([s.name for s in disciplines], ["Internal Finishes", "Stairs"])
        self.assertEqual(disciplines[0].scope_type, ProjectScope.ScopeType.PHASE)

        # The real Activity ID is used as the leaf's own code — not a
        # synthetic differentiator like the old assumed scheme's "NU001".
        seal = Activity.objects.get(project=project, code="MN6-A6-01-01")
        self.assertEqual(seal.name, "Seal")
        self.assertEqual(float(seal.progress_percent), 50.0)

        # Group nodes carry no Start/Finish of their own — rolled up from activities.
        self.assertEqual(stage.planned_start, d(2026, 1, 1))
        self.assertEqual(stage.planned_finish, d(2026, 2, 8))

    def test_reads_baseline_actual_duration_spi_and_schedule_variance(self):
        """Four columns the real export carries but nothing used to read:
        "BL Project Duration", "Actual Duration", "Schedule Performance
        Index", "Schedule Variance" — now captured on the Activity."""
        d = datetime.date
        header = ["Planex Code", "Activity ID", "Activity Name", "BL Project Duration",
                  "Original Duration", "Actual Duration", "Remaining Duration", "Start", "Finish",
                  "Total Float", "Activity % Complete", "Performance % Complete", "Schedule % Complete",
                  "Schedule Performance Index", "Budgeted Total Cost", "Earned Value Cost", "Schedule Variance"]
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1", "MN6-A6-01-01", "Seal",
             12, 12, 12, 0, d(2026, 1, 1), d(2026, 1, 10), None, 1, 1, 1, 1.02, 1000, 998, -2.5],
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        import_workbook(project, buf, source="planex_code.xlsx")

        seal = Activity.objects.get(project=project, code="MN6-A6-01-01")
        self.assertEqual(seal.baseline_duration, 12)
        self.assertEqual(seal.actual_duration, 12)
        self.assertAlmostEqual(float(seal.schedule_performance_index), 1.02)
        self.assertAlmostEqual(float(seal.schedule_variance), -2.5)

    def test_scope_label_read_from_the_wbs_heading_text(self):
        """Alongside its own code segment, the file's real WBS heading rows
        (interleaved with the coded leaf rows, indentation-based, the same
        as the old parser reads) carry a human-readable name. Each scope
        node should pick that up as `label`, keeping `name` as the stable
        code used for matching — a naming change or a missing heading on a
        later import must never affect `name`, only cosmetically update
        `label` (or leave it blank)."""
        d = datetime.date
        rows = [
            [None, "Mansora 6 - Revised Final", None, 0, None, None, None, 0, 0],
            # The "CON" tag's own WBS heading — real in the file, but dropped
            # from the code path as a placeholder (see segment_path). Sits at
            # the front, so it must not shift the backward alignment of the
            # levels that DO have a code counterpart.
            [None, "  Execution Phase", None, 0, None, None, None, 0, 0],
            [None, "    المرحلة الاولي (75 عمارة)", None, 0, None, None, None, 0, 0],
            [None, "      Zone (A)", None, 0, None, None, None, 0, 0],
            [None, "        (A6) Building", None, 0, None, None, None, 0, 0],
            [None, "          التشطيب الداخلي", None, 0, None, None, None, 0, 0],
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1", "A1", "Seal",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        import_workbook(project, self._workbook(rows), source="planex_code.xlsx")

        phase = ProjectScope.objects.get(project=project, name="PH1")
        self.assertEqual(phase.label, "المرحلة الاولي (75 عمارة)")

        zone = ProjectScope.objects.get(project=project, name="Z(A)")
        self.assertEqual(zone.label, "Zone (A)")

        building = ProjectScope.objects.get(project=project, name="Building 6")
        self.assertEqual(building.label, "(A6) Building")

        discipline = ProjectScope.objects.get(project=project, name="Internal Finishes")
        self.assertEqual(discipline.label, "التشطيب الداخلي")

    def test_scope_label_blank_when_no_wbs_heading_rows_exist(self):
        """A file with only coded leaf rows and no separate WBS heading rows
        (the shape every other test in this file uses) must import exactly
        as before — `label` stays blank, `name` (the code) is what's used,
        no crash from a missing heading stack entry."""
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 6-Internal Finishes-1", "A1", "Seal",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        import_workbook(project, self._workbook(rows), source="planex_code.xlsx")
        phase = ProjectScope.objects.get(project=project, name="PH1")
        self.assertEqual(phase.label, "")

    def test_budgeted_total_cost_header_drives_weighting(self):
        """This file's cost column is named "Budgeted Total Cost", not the
        original template's "Budgeted Material Cost" — must still be read
        for cost-based roll-up weighting."""
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-1", "A1", "Big", 1,
             d(2026, 1, 1), d(2026, 1, 2), 1, 9000, 9000],
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-2", "A2", "Small", 1,
             d(2026, 1, 1), d(2026, 1, 2), 0, 1000, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        result = import_workbook(project, self._workbook(rows), source="planex_code.xlsx")
        self.assertEqual(result["weighted_by"], "budget")

    def test_bare_tag_placeholders_all_the_way_falls_back_to_uncategorized(self):
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-0-0-0-0-DEC-1", "A1", "Orphan task",
             1, d(2026, 1, 1), d(2026, 1, 2), 0, 0, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        import_workbook(project, self._workbook(rows), source="planex_code.xlsx")

        # Every middle segment is a placeholder ("0" or "DEC" alone) -> no
        # real tree path -> this parser doesn't match the row at all, so
        # detection falls through (no Planex Code sheet actually matched).
        self.assertFalse(ProjectScope.objects.filter(project=project).exists())

    def test_key_milestones_branch_has_no_planex_code_but_still_imports(self):
        """A real P6 export keeps its key dates (project start/finish,
        handover milestones) as zero-work activities under one WBS heading
        — carrying NO Planex Code at all, since they aren't coded discipline
        work. The code-driven walk must not just silently drop them; they
        belong in the Milestones panel, same as the old indentation
        parser's own milestone branch."""
        d = datetime.date
        rows = [
            # Planex-Code-driven real work, so the sheet is detected at all.
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-1", "A1", "Wiring",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
            # The milestone WBS branch — no Planex Code on any of these rows.
            [None, "  Key Milestones", None, 0, d(2026, 1, 1), d(2026, 12, 1), None, 0, 0],
            [None, "MS-START", "Project start", 0, d(2026, 1, 1), None, 1, 0, 0],
            [None, "MS-END", "Project end", 0, None, d(2026, 12, 1), 0, 0, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        result = import_workbook(project, self._workbook(rows), source="planex_code.xlsx")

        self.assertEqual(result["activities"], 1)  # only the coded row
        self.assertEqual(result["milestones"], 2)
        from .models import Milestone
        titles = set(Milestone.objects.filter(project=project).values_list("title", flat=True))
        self.assertEqual(titles, {"Project start", "Project end"})
        start_ms = Milestone.objects.get(project=project, title="Project start")
        self.assertEqual(start_ms.status, Milestone.Status.COMPLETED)  # pct=1 -> 100%
        self.assertEqual(start_ms.progress_percent, 100)
        end_ms = Milestone.objects.get(project=project, title="Project end")
        self.assertEqual(end_ms.status, Milestone.Status.UPCOMING)  # pct=0
        # A real 0% (not missing) must still be stored, not left null.
        self.assertEqual(end_ms.progress_percent, 0)

        # The milestone branch must not also appear as a schedule scope.
        self.assertFalse(ProjectScope.objects.filter(project=project, name="Key Milestones").exists())

    def test_milestone_with_no_pct_column_value_stores_null_not_zero(self):
        """A milestone row with nothing at all in the "Activity % Complete"
        column must store progress_percent=None — never silently coerced to
        0, which would be indistinguishable from a real "0% complete"."""
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-1", "A1", "Wiring",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
            [None, "  Key Milestones", None, 0, d(2026, 1, 1), d(2026, 12, 1), None, 0, 0],
            [None, "MS-START", "Project start", 0, d(2026, 1, 1), None, None, 0, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        import_workbook(project, self._workbook(rows), source="planex_code.xlsx")
        from .models import Milestone
        ms = Milestone.objects.get(project=project, title="Project start")
        self.assertIsNone(ms.progress_percent)
        self.assertEqual(ms.status, Milestone.Status.UPCOMING)  # still defaults sensibly

    def test_ms_coded_milestone_resolves_scope_to_the_matching_building(self):
        """The agreed convention for once the team codes the Key Milestones
        branch: "MS" as the 2nd segment, with the same zone/building
        segments as the discipline rows when the milestone is tied to one
        building. It should resolve to the SAME ProjectScope the discipline
        rows created — not a raw-text match, a real FK."""
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-1", "A1", "Wiring",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
            # Tied to that same building — should resolve to its scope.
            ["MN(6)-MS-PH1-Z(A)-Building 1-1", "MS1", "(A1) Building handover",
             0, None, d(2026, 3, 1), 1, 0, 0],
            # Project-wide — no zone/building segments, scope stays null.
            ["MN(6)-MS-2", "MS2", "Project start",
             0, d(2026, 1, 1), None, 1, 0, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        result = import_workbook(project, self._workbook(rows), source="planex_code.xlsx")

        self.assertEqual(result["activities"], 1)
        self.assertEqual(result["milestones"], 2)

        from .models import Milestone
        building = ProjectScope.objects.get(project=project, name="Building 1")
        handover = Milestone.objects.get(project=project, title="(A1) Building handover")
        self.assertEqual(handover.scope_id, building.id)
        self.assertEqual(handover.status, Milestone.Status.COMPLETED)

        project_start = Milestone.objects.get(project=project, title="Project start")
        self.assertIsNone(project_start.scope_id)

    def test_ms_coded_milestones_take_priority_over_indentation_fallback(self):
        """When the sheet has real MS-coded milestone rows, use those instead
        of the indentation/keyword fallback — even if a WBS heading elsewhere
        happens to look milestone-ish, it should be ignored in favor of the
        explicit code."""
        d = datetime.date
        rows = [
            ["MN(6)-CON-0-0-PH1-Z(A)-0-Building 1-ELEC-1", "A1", "Wiring",
             1, d(2026, 1, 1), d(2026, 1, 2), 1, 1000, 1000],
            ["MN(6)-MS-1", "MS1", "Project start",
             0, d(2026, 1, 1), None, 1, 0, 0],
        ]
        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        result = import_workbook(project, self._workbook(rows), source="planex_code.xlsx")
        self.assertEqual(result["milestones"], 1)
        from .models import Milestone
        self.assertEqual(
            set(Milestone.objects.filter(project=project).values_list("title", flat=True)),
            {"Project start"},
        )

    def test_no_planex_code_column_falls_through_to_old_parser(self):
        """A file with no "Planex Code" column at all (the previous
        template shape) must still import via the leading-space scheme,
        completely unaffected by this module."""
        d = datetime.date
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Activity ID", "Activity Name", "Original Duration", "Start", "Finish",
                  "Activity % Complete", "Budgeted Material Cost", "Earned Value Cost"])
        ws.append(["  Construction Phase", None, 0, d(2026, 1, 1), d(2026, 2, 1), None, 0, 0])
        ws.append(["CN.01", "Foundation", 10, d(2026, 1, 1), d(2026, 1, 10), 0.5, 1000, 500])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        result = import_workbook(project, buf, source="legacy.xlsx")
        self.assertEqual(result["source_kind"], "p6_schedule")
        self.assertTrue(ProjectScope.objects.filter(project=project, name="Construction Phase").exists())
