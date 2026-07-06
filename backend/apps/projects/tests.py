"""Project API tests: tenant isolation, permission enforcement, CRUD, archive."""
from django.test import SimpleTestCase, TestCase
from django.urls import reverse

from apps.accounts.constants import COMPANY_ADMIN_PERMISSIONS, Permission, SeededRole
from apps.accounts.models import Company, Membership, Role, User
from .imports import _guess_discipline, import_schedule, parse_sheet
from .models import Project


class ImportParserTests(SimpleTestCase):
    """The zone-sheet parser detects the grid (with or without a weight column)
    and averages each task across its subzones."""

    def test_detects_grid_with_weight_column(self):
        rows = [
            (None, None, 1, 2, 3),
            (None, None, "(A1)", "(A2)", "(A3)"),
            ("W", "Phase 1", 0.5, 0.5, 0.5),   # phase header (col-A "W") -> skipped
            (2.0, "Task1", 1, 1, 1),           # cells 100/100/100, weight 2
            (1.0, "Task2", 0, 0.5, 1),         # cells 0/50/100, weight 1
        ]
        sheet = parse_sheet(rows)
        self.assertEqual(sheet["subzones"], ["(A1)", "(A2)", "(A3)"])
        self.assertEqual(len(sheet["tasks"]), 2)
        self.assertEqual(sheet["tasks"][0]["weight"], 2.0)
        self.assertEqual(sheet["tasks"][0]["phase"], "Phase 1")
        self.assertEqual(sheet["tasks"][0]["cells"], [100.0, 100.0, 100.0])
        self.assertEqual(sheet["tasks"][1]["cells"], [0.0, 50.0, 100.0])

    def test_detects_grid_without_weight_column(self):
        # Like ZONE (B): codes start in column 1, name is column 0, no weight col.
        rows = [
            (None, 1, 2),
            (None, "(B1)", "(B2)"),
            ("summary", 0.4, 0.4),         # first row is the summary -> skipped
            ("Task1", 1, 0),              # cells 100/0, default weight 1
        ]
        sheet = parse_sheet(rows)
        self.assertEqual(sheet["subzones"], ["(B1)", "(B2)"])
        self.assertEqual(len(sheet["tasks"]), 1)
        self.assertEqual(sheet["tasks"][0]["weight"], 1.0)
        self.assertEqual(sheet["tasks"][0]["cells"], [100.0, 0.0])

    def test_guesses_discipline_from_phase_name(self):
        self.assertEqual(_guess_discipline("الاعمال الكهربائية"), "electrical")
        self.assertEqual(_guess_discipline("اعمال الخرسانة"), "concrete")
        self.assertEqual(_guess_discipline("اعمال التشطيبات"), "architecture")
        self.assertEqual(_guess_discipline("شبكات الصرف الصحي"), "mechanical")
        self.assertEqual(_guess_discipline("Random unrelated text"), "")


class ScheduleImportTests(TestCase):
    """`import_schedule` matches a flat Activity Name/Start/Finish export (the
    shape Primavera P6 exports to Excel) to existing scopes by name and sets
    their dates — never touches structure."""

    def _workbook(self, header, rows):
        import io

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(header)
        for row in rows:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_matches_by_name_and_sets_dates(self):
        import datetime

        from apps.accounts.models import Company
        from .models import ProjectScope

        company = Company.objects.create(name="Acme")
        project = Project.objects.create(company=company, name="Tower", project_type="commercial")
        zone = ProjectScope.objects.create(company=company, project=project, scope_type="zone", name="ZONE (A)")

        wb = self._workbook(
            ["Activity Name", "Start", "Finish"],
            [["ZONE (A)", datetime.date(2026, 1, 1), datetime.date(2026, 6, 1)],
             ["No Match Here", datetime.date(2026, 1, 1), datetime.date(2026, 2, 1)]],
        )
        result = import_schedule(project, wb)
        self.assertEqual(result, {"matched": 1, "unmatched": 1, "total_rows": 2})
        zone.refresh_from_db()
        self.assertEqual(zone.planned_start, datetime.date(2026, 1, 1))
        self.assertEqual(zone.planned_finish, datetime.date(2026, 6, 1))

STRONG_PW = "Str0ngPassw0rd!"


class ProjectApiTests(TestCase):
    def setUp(self):
        self.company_a = Company.objects.create(name="Acme")
        self.admin_role = Role.objects.create(
            company=self.company_a, name=SeededRole.COMPANY_ADMIN,
            permissions=COMPANY_ADMIN_PERMISSIONS)
        self.admin = User.objects.create_user(
            email="admin@acme.com", password=STRONG_PW, company=self.company_a)
        Membership.objects.create(company=self.company_a, user=self.admin, role=self.admin_role)

        self.viewer_role = Role.objects.create(
            company=self.company_a, name="Viewer",
            permissions=[Permission.VIEW_PROJECTS.value, Permission.VIEW_SCHEDULE.value])
        self.viewer = User.objects.create_user(
            email="viewer@acme.com", password=STRONG_PW, company=self.company_a)
        Membership.objects.create(company=self.company_a, user=self.viewer, role=self.viewer_role)

        # Other tenant + its project + user (must stay invisible to company A).
        self.company_b = Company.objects.create(name="Globex")
        self.project_b = Project.objects.create(
            company=self.company_b, name="B Tower", project_type="commercial")
        self.user_b = User.objects.create_user(
            email="bob@globex.com", password=STRONG_PW, company=self.company_b)

    def login(self, email):
        resp = self.client.post(
            reverse("auth-login"), {"email": email, "password": STRONG_PW},
            content_type="application/json")
        self.assertEqual(resp.status_code, 200, resp.content)

    def test_create_requires_manage_projects(self):
        self.login("viewer@acme.com")
        resp = self.client.post(
            "/api/projects/", {"name": "Mall", "project_type": "commercial"},
            content_type="application/json")
        self.assertEqual(resp.status_code, 403)

    def test_admin_creates_and_lists_project(self):
        self.login("admin@acme.com")
        resp = self.client.post(
            "/api/projects/",
            {"name": "Mall", "project_type": "commercial", "location": "Dubai"},
            content_type="application/json")
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertEqual(resp.json()["project_type_display"], "Commercial")

        names = {p["name"] for p in self.client.get("/api/projects/").json()["results"]}
        self.assertIn("Mall", names)
        self.assertNotIn("B Tower", names)  # company B isolated

    def test_viewer_can_read(self):
        Project.objects.create(company=self.company_a, name="Villa", project_type="residential")
        self.login("viewer@acme.com")
        self.assertEqual(self.client.get("/api/projects/").status_code, 200)

    def test_cannot_access_other_company_project(self):
        self.login("admin@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{self.project_b.id}/").status_code, 404)

    def test_archive_and_filter(self):
        p = Project.objects.create(company=self.company_a, name="Depot", project_type="industrial")
        self.login("admin@acme.com")
        self.client.patch(f"/api/projects/{p.id}/", {"is_archived": True},
                          content_type="application/json")
        active = {x["name"] for x in self.client.get("/api/projects/?status=active").json()["results"]}
        archived = {x["name"] for x in self.client.get("/api/projects/?status=archived").json()["results"]}
        self.assertNotIn("Depot", active)
        self.assertIn("Depot", archived)

    def test_duplicate_name_rejected(self):
        Project.objects.create(company=self.company_a, name="Tower", project_type="commercial")
        self.login("admin@acme.com")
        resp = self.client.post("/api/projects/", {"name": "Tower", "project_type": "commercial"},
                                content_type="application/json")
        self.assertEqual(resp.status_code, 400)

    # ── Work hierarchy ────────────────────────────────────────────────────
    def test_build_structure_and_rollup(self):
        p = Project.objects.create(company=self.company_a, name="Hospital", project_type="commercial")
        self.login("admin@acme.com")
        base = f"/api/projects/{p.id}"
        # phase -> activity x2 with different weights/progress
        phase = self.client.post(f"{base}/scopes/", {"scope_type": "phase", "name": "Substructure"},
                                 content_type="application/json").json()
        a1 = self.client.post(f"{base}/activities/",
                              {"scope": phase["id"], "name": "Excavation", "weight": "3", "progress_percent": "100"},
                              content_type="application/json")
        a2 = self.client.post(f"{base}/activities/",
                              {"scope": phase["id"], "name": "Piling", "weight": "1", "progress_percent": "0"},
                              content_type="application/json")
        self.assertEqual(a1.status_code, 201, a1.content)
        self.assertEqual(a2.status_code, 201, a2.content)
        # overall = (100*3 + 0*1) / (3+1) = 75
        struct = self.client.get(f"{base}/structure/").json()
        self.assertEqual(struct["overall_progress"], 75.0)
        self.assertEqual(len(struct["scopes"]), 1)
        self.assertEqual(struct["activity_count"], 2)
        self.assertEqual(struct["scope_activity_counts"][phase["id"]], 2)
        # activities are lazy-loaded per scope
        acts = self.client.get(f"{base}/scopes/{phase['id']}/activities/").json()
        self.assertEqual(len(acts), 2)

    def test_scope_dates_and_discipline_round_trip(self):
        p = Project.objects.create(company=self.company_a, name="Resort3", project_type="commercial")
        self.login("admin@acme.com")
        base = f"/api/projects/{p.id}"
        resp = self.client.post(f"{base}/scopes/", {
            "scope_type": "phase", "name": "Electrical works", "discipline": "electrical",
            "planned_start": "2026-01-01", "planned_finish": "2026-03-01",
        }, content_type="application/json")
        self.assertEqual(resp.status_code, 201, resp.content)
        body = resp.json()
        self.assertEqual(body["discipline"], "electrical")
        self.assertEqual(body["discipline_display"], "Electrical")
        self.assertEqual(body["planned_start"], "2026-01-01")
        self.assertEqual(body["planned_finish"], "2026-03-01")

    def test_update_activity_progress_recomputes(self):
        p = Project.objects.create(company=self.company_a, name="Bridge", project_type="infrastructure")
        self.login("admin@acme.com")
        base = f"/api/projects/{p.id}"
        phase = self.client.post(f"{base}/scopes/", {"scope_type": "phase", "name": "Deck"},
                                 content_type="application/json").json()
        act = self.client.post(f"{base}/activities/",
                               {"scope": phase["id"], "name": "Pour", "weight": "1", "progress_percent": "0"},
                               content_type="application/json").json()
        self.client.patch(f"{base}/activities/{act['id']}/", {"progress_percent": "50"},
                          content_type="application/json")
        detail = self.client.get(f"{base}/").json()
        self.assertEqual(detail["overall_progress"], 50.0)
        self.assertEqual(detail["activity_count"], 1)

    def test_viewer_cannot_edit_structure(self):
        p = Project.objects.create(company=self.company_a, name="Depot2", project_type="industrial")
        self.login("viewer@acme.com")
        resp = self.client.post(f"/api/projects/{p.id}/scopes/", {"scope_type": "phase", "name": "X"},
                                content_type="application/json")
        self.assertEqual(resp.status_code, 403)

    def test_cannot_build_structure_on_other_company_project(self):
        self.login("admin@acme.com")
        resp = self.client.post(f"/api/projects/{self.project_b.id}/scopes/",
                                {"scope_type": "phase", "name": "X"}, content_type="application/json")
        self.assertEqual(resp.status_code, 404)

    # ── Team ──────────────────────────────────────────────────────────────
    def test_add_member_surfaces_company_role_on_detail(self):
        p = Project.objects.create(company=self.company_a, name="Clinic", project_type="commercial")
        self.login("admin@acme.com")
        resp = self.client.post(
            f"/api/projects/{p.id}/members/",
            {"user_ids": [str(self.viewer.id)]},
            content_type="application/json")
        self.assertEqual(resp.status_code, 201, resp.content)
        # The member's displayed role is their real company role — read-only here.
        self.assertEqual(resp.json()[0]["role_names"], ["Viewer"])
        detail = self.client.get(f"/api/projects/{p.id}/").json()
        self.assertEqual(detail["team_count"], 1)

    def test_cannot_add_user_from_other_company(self):
        p = Project.objects.create(company=self.company_a, name="Bridge2", project_type="infrastructure")
        self.login("admin@acme.com")
        resp = self.client.post(
            f"/api/projects/{p.id}/members/", {"user_ids": [str(self.user_b.id)]},
            content_type="application/json")
        self.assertEqual(resp.status_code, 400)

    def test_viewer_cannot_add_member(self):
        p = Project.objects.create(company=self.company_a, name="Depot3", project_type="industrial")
        self.login("viewer@acme.com")
        resp = self.client.post(
            f"/api/projects/{p.id}/members/", {"user_ids": [str(self.admin.id)]},
            content_type="application/json")
        self.assertEqual(resp.status_code, 403)

    # ── Scope access ──────────────────────────────────────────────────────
    def test_scope_access_restricts_visible_zones(self):
        from .models import ProjectScope, ProjectScopeAccess
        p = Project.objects.create(company=self.company_a, name="Mall2", project_type="commercial")
        z1 = ProjectScope.objects.create(company=self.company_a, project=p, scope_type="zone", name="Z1")
        z2 = ProjectScope.objects.create(company=self.company_a, project=p, scope_type="zone", name="Z2")
        ProjectScope.objects.create(company=self.company_a, project=p, parent=z1, scope_type="area", name="Z1-A")

        # viewer has only VIEW_PROJECTS -> can be restricted
        self.login("viewer@acme.com")
        names = {s["name"] for s in self.client.get(f"/api/projects/{p.id}/structure/").json()["scopes"]}
        self.assertEqual(names, {"Z1", "Z2", "Z1-A"})  # no grants -> full

        ProjectScopeAccess.objects.create(company=self.company_a, project=p, user=self.viewer, scope=z1)
        names = {s["name"] for s in self.client.get(f"/api/projects/{p.id}/structure/").json()["scopes"]}
        self.assertEqual(names, {"Z1", "Z1-A"})  # restricted to Z1 + descendants

        # ...but the admin (MANAGE_PROJECTS) still sees everything.
        self.login("admin@acme.com")
        names = {s["name"] for s in self.client.get(f"/api/projects/{p.id}/structure/").json()["scopes"]}
        self.assertEqual(names, {"Z1", "Z2", "Z1-A"})

    def test_manage_member_scope_access(self):
        from .models import ProjectScope
        p = Project.objects.create(company=self.company_a, name="Mall3", project_type="commercial")
        z1 = ProjectScope.objects.create(company=self.company_a, project=p, scope_type="zone", name="Z1")
        self.login("admin@acme.com")
        m = self.client.post(f"/api/projects/{p.id}/members/",
                             {"user_ids": [str(self.viewer.id)]},
                             content_type="application/json").json()[0]
        put = self.client.put(f"/api/projects/{p.id}/members/{m['id']}/scope-access/",
                              {"scope_ids": [str(z1.id)]}, content_type="application/json")
        self.assertEqual(put.status_code, 200, put.content)
        self.assertEqual(put.json()["scope_ids"], [str(z1.id)])
        got = self.client.get(f"/api/projects/{p.id}/members/{m['id']}/scope-access/").json()
        self.assertEqual(got["scope_ids"], [str(z1.id)])

    def test_add_members_batch_with_scope(self):
        # Adding a member accepts a scope grant in the same call.
        from .models import ProjectScope, ProjectScopeAccess
        p = Project.objects.create(company=self.company_a, name="Batch1", project_type="commercial")
        z1 = ProjectScope.objects.create(company=self.company_a, project=p, scope_type="zone", name="Z1")
        self.login("admin@acme.com")
        resp = self.client.post(f"/api/projects/{p.id}/members/", {
            "user_ids": [str(self.viewer.id)], "scope_ids": [str(z1.id)],
        }, content_type="application/json")
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertTrue(ProjectScopeAccess.objects.filter(project=p, user=self.viewer, scope=z1).exists())

    # ── Approval chain ────────────────────────────────────────────────────
    def _activity(self):
        p = Project.objects.create(company=self.company_a, name="Lab", project_type="commercial")
        from .models import Activity, ProjectScope
        zone = ProjectScope.objects.create(company=self.company_a, project=p, scope_type="zone", name="Z")
        return p, Activity.objects.create(company=self.company_a, project=p, scope=zone, name="Pour",
                                          weight=1, progress_percent=0)

    def _grant(self, user, *perms):
        role = Role.objects.create(company=self.company_a, name="R-" + "-".join(perms), permissions=list(perms))
        Membership.objects.create(company=self.company_a, user=user, role=role)

    def test_full_approval_chain_updates_activity(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        # admin already has all company perms (review + approve)
        self.login("eng@acme.com")
        r = self.client.post(f"/api/projects/{p.id}/submissions/",
                             {"activity": str(act.id), "submitted_progress": "60"}, content_type="application/json")
        self.assertEqual(r.status_code, 201, r.content)
        sid = r.json()["id"]
        self.assertEqual(r.json()["status"], "pending_review")

        self.login("admin@acme.com")
        rev = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                               {"decision": "approve"}, content_type="application/json")
        self.assertEqual(rev.json()["status"], "pending_pm")
        app = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/approve/",
                               {"decision": "approve"}, content_type="application/json")
        self.assertEqual(app.json()["status"], "accepted")
        act.refresh_from_db()
        self.assertEqual(float(act.progress_percent), 60.0)  # accepted -> official

    def test_audit_timestamps_recorded(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng3@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login("eng3@acme.com")
        sid = self.client.post(f"/api/projects/{p.id}/submissions/",
                               {"activity": str(act.id), "submitted_progress": "30"},
                               content_type="application/json").json()["id"]
        self.login("admin@acme.com")
        rev = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                               {"decision": "approve"}, content_type="application/json").json()
        self.assertIsNotNone(rev["reviewed_at"])
        self.assertIsNone(rev["decided_at"])
        app = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/approve/",
                               {"decision": "approve"}, content_type="application/json").json()
        self.assertIsNotNone(app["decided_at"])

    # ── Submission photos ─────────────────────────────────────────────────
    # Minimal valid 1x1 PNG, used across these tests as upload fixture bytes.
    _PNG_BYTES = (
        b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00'
        b'\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4\x00\x00\x00\x00IEND\xaeB`\x82'
    )

    def _png(self, name="photo.png"):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile(name, self._PNG_BYTES, content_type="image/png")

    def test_submission_photo_upload_and_serialized(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng4@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login("eng4@acme.com")
        sid = self.client.post(f"/api/projects/{p.id}/submissions/",
                               {"activity": str(act.id), "submitted_progress": "40"},
                               content_type="application/json").json()["id"]

        resp = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/images/",
                                {"image": self._png(), "caption": "Poured slab"})
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertEqual(resp.json()["caption"], "Poured slab")

        # eng4 only has SUBMIT_PROGRESS (not VIEW_PROJECTS) -- check the listing as
        # someone who can actually view the project.
        self.login("admin@acme.com")
        listing = self.client.get(f"/api/projects/{p.id}/submissions/").json()
        sub = next(s for s in listing if s["id"] == sid)
        self.assertEqual(len(sub["images"]), 1)
        self.assertEqual(sub["images"][0]["caption"], "Poured slab")

        # The file streams back through the authed endpoint.
        img_id = resp.json()["id"]
        file_resp = self.client.get(f"/api/projects/{p.id}/submissions/{sid}/images/{img_id}/file/")
        self.assertEqual(file_resp.status_code, 200)

    def test_submission_photo_requires_submit_permission(self):
        p, act = self._activity()
        sid = self._submit(p, act)  # creates the submission via a throwaway engineer
        # A plain viewer (no submit_progress) cannot attach a photo.
        self.login("viewer@acme.com")
        resp = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/images/",
                                {"image": self._png()})
        self.assertEqual(resp.status_code, 403)

    def test_submission_photo_rejects_bad_content_type(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng5@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login("eng5@acme.com")
        sid = self.client.post(f"/api/projects/{p.id}/submissions/",
                               {"activity": str(act.id), "submitted_progress": "10"},
                               content_type="application/json").json()["id"]
        from django.core.files.uploadedfile import SimpleUploadedFile
        bad = SimpleUploadedFile("notes.txt", b"hello", content_type="text/plain")
        resp = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/images/", {"image": bad})
        self.assertEqual(resp.status_code, 400)

    def test_accepted_submission_photo_appears_in_progress_gallery(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng6@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login("eng6@acme.com")
        sid = self.client.post(f"/api/projects/{p.id}/submissions/",
                               {"activity": str(act.id), "submitted_progress": "55"},
                               content_type="application/json").json()["id"]
        self.client.post(f"/api/projects/{p.id}/submissions/{sid}/images/",
                         {"image": self._png(), "caption": "Rebar laid"})

        self.login("admin@acme.com")
        self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                         {"decision": "approve"}, content_type="application/json")
        self.client.post(f"/api/projects/{p.id}/submissions/{sid}/approve/",
                         {"decision": "approve"}, content_type="application/json")

        gallery = self.client.get(f"/api/projects/{p.id}/progress-images/").json()
        matches = [g for g in gallery if g["activity_id"] == str(act.id)]
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0]["caption"], "Rebar laid")

    def test_backfill_migration_carries_old_accepted_photos(self):
        # Simulate "old data": an accepted submission with a photo but no
        # ProgressEntry (as it would exist before the carry-over feature).
        import importlib
        from django.apps import apps as global_apps
        from django.utils import timezone
        from .models import ProgressEntry, ProgressImage, ProgressSubmission, SubmissionImage

        p, act = self._activity()
        sub = ProgressSubmission.objects.create(
            company=self.company_a, project=p, activity=act,
            submitted_by=self.admin, previous_progress=0, submitted_progress=70,
            status=ProgressSubmission.Status.ACCEPTED, note="old one",
            decided_at=timezone.now(),
        )
        SubmissionImage.objects.create(
            company=self.company_a, submission=sub, image=self._png(),
            caption="Old evidence", uploaded_by=self.admin)
        self.assertFalse(ProgressEntry.objects.filter(activity=act).exists())

        mig = importlib.import_module(
            "apps.projects.migrations.0025_backfill_accepted_submission_photos")
        mig.backfill(global_apps, None)

        gallery = ProgressImage.objects.filter(entry__activity=act)
        self.assertEqual(gallery.count(), 1)
        self.assertEqual(gallery.first().caption, "Old evidence")

        # Idempotent: a second run must not duplicate.
        mig.backfill(global_apps, None)
        self.assertEqual(ProgressImage.objects.filter(entry__activity=act).count(), 1)

    def test_reject_requires_comment_and_keeps_progress(self):
        p, act = self._activity()
        from apps.accounts.constants import Permission as P
        eng = User.objects.create_user(email="eng2@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login("eng2@acme.com")
        sid = self.client.post(f"/api/projects/{p.id}/submissions/",
                               {"activity": str(act.id), "submitted_progress": "50"},
                               content_type="application/json").json()["id"]
        self.login("admin@acme.com")
        no_comment = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                                      {"decision": "reject"}, content_type="application/json")
        self.assertEqual(no_comment.status_code, 400)
        ok = self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                              {"decision": "reject", "comment": "Recheck zone B"}, content_type="application/json")
        self.assertEqual(ok.json()["status"], "reviewer_rejected")
        act.refresh_from_db()
        self.assertEqual(float(act.progress_percent), 0.0)  # unchanged

    def test_viewer_cannot_submit(self):
        p, act = self._activity()
        self.login("viewer@acme.com")
        r = self.client.post(f"/api/projects/{p.id}/submissions/",
                             {"activity": str(act.id), "submitted_progress": "10"}, content_type="application/json")
        self.assertEqual(r.status_code, 403)

    # Shared helper: create an engineer, submit progress, return the submission id.
    def _submit(self, project, activity, value="40"):
        eng_email = f"eng-{activity.id}@acme.com"
        eng = User.objects.create_user(email=eng_email, password=STRONG_PW, company=self.company_a)
        from apps.accounts.constants import Permission as P
        self._grant(eng, P.SUBMIT_PROGRESS.value)
        self.login(eng_email)
        return self.client.post(f"/api/projects/{project.id}/submissions/",
                                {"activity": str(activity.id), "submitted_progress": value},
                                content_type="application/json").json()["id"]

    # ── P6 export ─────────────────────────────────────────────────────────
    def test_p6_export_returns_xlsx(self):
        p, act = self._activity()
        self.login("admin@acme.com")  # holds export_reports via company-admin perms
        resp = self.client.get(f"/api/projects/{p.id}/export/p6/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertTrue(resp["Content-Disposition"].endswith('.xlsx"'))
        self.assertEqual(resp.content[:2], b"PK")  # xlsx is a zip

    def test_p6_prepare_instant_without_source_workbook(self):
        p, _ = self._activity()
        self.login("admin@acme.com")
        resp = self.client.post(f"/api/projects/{p.id}/export/p6/prepare/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["mode"], "instant")

    def test_p6_export_forbidden_without_perm(self):
        p, _ = self._activity()
        # A user with only SUBMIT_PROGRESS (no export/manage).
        from apps.accounts.constants import Permission as P
        u = User.objects.create_user(email="noexp@acme.com", password=STRONG_PW, company=self.company_a)
        self._grant(u, P.SUBMIT_PROGRESS.value)
        self.login("noexp@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{p.id}/export/p6/").status_code, 403)

    # ── Global search ─────────────────────────────────────────────────────
    def test_search_finds_projects_and_activities(self):
        p, act = self._activity()  # project "Lab", activity "Pour"
        self.login("viewer@acme.com")  # has VIEW_PROJECTS
        data = self.client.get("/api/search/?q=La").json()
        self.assertIn("Lab", [pr["name"] for pr in data["projects"]])
        data2 = self.client.get("/api/search/?q=Pour").json()
        self.assertIn("Pour", [a["name"] for a in data2["activities"]])
        self.assertEqual(data2["activities"][0]["project_name"], "Lab")

    def test_search_short_query_is_empty(self):
        self._activity()
        self.login("viewer@acme.com")
        data = self.client.get("/api/search/?q=L").json()
        self.assertEqual(data["projects"], [])
        self.assertEqual(data["activities"], [])

    # ── Notifications ─────────────────────────────────────────────────────
    def test_submit_notifies_reviewer(self):
        p, act = self._activity()
        self._submit(p, act)  # leaves client logged in as the engineer
        self.login("admin@acme.com")  # admin can review
        data = self.client.get("/api/notifications/").json()
        self.assertGreaterEqual(data["unread_count"], 1)
        self.assertIn("submitted", [n["kind"] for n in data["results"]])

    def test_reject_notifies_submitter_and_mark_read(self):
        p, act = self._activity()
        sid = self._submit(p, act)
        eng_email = f"eng-{act.id}@acme.com"
        self.login("admin@acme.com")
        self.client.post(f"/api/projects/{p.id}/submissions/{sid}/review/",
                         {"decision": "reject", "comment": "Recheck"}, content_type="application/json")
        self.login(eng_email)
        data = self.client.get("/api/notifications/").json()
        self.assertEqual(data["unread_count"], 1)
        self.assertEqual(data["results"][0]["kind"], "review_rejected")
        read = self.client.post("/api/notifications/read/")
        self.assertEqual(read.json()["unread_count"], 0)

    def test_notifications_are_per_recipient(self):
        # The submitter shouldn't receive the "submitted" notice meant for reviewers.
        p, act = self._activity()
        self._submit(p, act)  # client is the engineer
        data = self.client.get("/api/notifications/").json()
        self.assertEqual(data["unread_count"], 0)

    def test_notification_filter_and_per_item_read(self):
        p, act = self._activity()
        self._submit(p, act)
        self.login("admin@acme.com")
        unread = self.client.get("/api/notifications/?filter=unread").json()
        self.assertGreaterEqual(len(unread["results"]), 1)
        nid = unread["results"][0]["id"]
        # Mark just this one read.
        self.client.post("/api/notifications/read/", {"ids": [nid]}, content_type="application/json")
        after_unread = [n["id"] for n in self.client.get("/api/notifications/?filter=unread").json()["results"]]
        after_read = [n["id"] for n in self.client.get("/api/notifications/?filter=read").json()["results"]]
        self.assertNotIn(nid, after_unread)
        self.assertIn(nid, after_read)

    # ── Milestones ────────────────────────────────────────────────────────
    def test_milestones_crud_and_permissions(self):
        p = Project.objects.create(company=self.company_a, name="Resort", project_type="commercial")
        self.login("admin@acme.com")
        resp = self.client.post(
            f"/api/projects/{p.id}/milestones/",
            {"title": "Kickoff", "date": "2026-01-10", "status": "completed"},
            content_type="application/json")
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertEqual(resp.json()["status_display"], "Completed")
        mid = resp.json()["id"]
        self.assertEqual(len(self.client.get(f"/api/projects/{p.id}/milestones/").json()), 1)
        self.assertEqual(self.client.delete(f"/api/projects/{p.id}/milestones/{mid}/").status_code, 204)

    def test_viewer_cannot_add_milestone(self):
        p = Project.objects.create(company=self.company_a, name="Resort2", project_type="commercial")
        self.login("viewer@acme.com")
        resp = self.client.post(f"/api/projects/{p.id}/milestones/", {"title": "X"},
                                content_type="application/json")
        self.assertEqual(resp.status_code, 403)

    def test_assignable_users_excludes_existing_members(self):
        p = Project.objects.create(company=self.company_a, name="Plaza", project_type="commercial")
        self.login("admin@acme.com")
        self.client.post(f"/api/projects/{p.id}/members/",
                         {"user_id": str(self.viewer.id)}, content_type="application/json")
        emails = {u["email"] for u in self.client.get(f"/api/projects/{p.id}/assignable-users/").json()}
        self.assertIn("admin@acme.com", emails)
        self.assertNotIn("viewer@acme.com", emails)  # already a member


class FinanceSubmittalApiTests(TestCase):
    """Cash flow + invoices are gated by the dedicated finance permissions;
    submittals by the dedicated submittal permissions — both company-role-based."""

    def setUp(self):
        self.company = Company.objects.create(name="Acme")
        admin_role = Role.objects.create(
            company=self.company, name=SeededRole.COMPANY_ADMIN, permissions=COMPANY_ADMIN_PERMISSIONS)
        self.admin = User.objects.create_user(email="fa@acme.com", password=STRONG_PW, company=self.company)
        Membership.objects.create(company=self.company, user=self.admin, role=admin_role)

        # Can view the project + finances, but cannot manage anything.
        fin_role = Role.objects.create(
            company=self.company, name="FinViewer",
            permissions=[Permission.VIEW_PROJECTS.value, Permission.VIEW_FINANCES.value])
        self.fin_viewer = User.objects.create_user(email="fv@acme.com", password=STRONG_PW, company=self.company)
        Membership.objects.create(company=self.company, user=self.fin_viewer, role=fin_role)

        # Can view/manage submittals, but has no finance permission at all.
        plain_role = Role.objects.create(
            company=self.company, name="Plain",
            permissions=[Permission.VIEW_PROJECTS.value, Permission.VIEW_SUBMITTALS.value,
                         Permission.MANAGE_SUBMITTALS.value])
        self.plain = User.objects.create_user(email="pl@acme.com", password=STRONG_PW, company=self.company)
        Membership.objects.create(company=self.company, user=self.plain, role=plain_role)

        self.project = Project.objects.create(company=self.company, name="Tower", project_type="commercial")

    def login(self, email):
        resp = self.client.post(reverse("auth-login"), {"email": email, "password": STRONG_PW},
                                content_type="application/json")
        self.assertEqual(resp.status_code, 200, resp.content)

    def test_finance_hidden_without_finances_perm(self):
        # plain member (submittals only) has no finance permission -> denied
        self.login("pl@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{self.project.id}/cashflow/").status_code, 403)
        self.assertEqual(self.client.get(f"/api/projects/{self.project.id}/invoices/").status_code, 403)

    def test_finance_viewer_can_read_not_write(self):
        self.login("fv@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{self.project.id}/cashflow/").status_code, 200)
        resp = self.client.put(f"/api/projects/{self.project.id}/cashflow/",
                               [{"month": "2026-01-15", "planned": 100, "actual": 80}],
                               content_type="application/json")
        self.assertEqual(resp.status_code, 403)  # needs MANAGE_FINANCES

    def test_cashflow_bulk_replace_normalises_month(self):
        self.login("fa@acme.com")
        resp = self.client.put(f"/api/projects/{self.project.id}/cashflow/",
                               [{"month": "2026-01-15", "planned": 100, "actual": 80},
                                {"month": "2026-02-01", "planned": 200, "actual": 150}],
                               content_type="application/json")
        self.assertEqual(resp.status_code, 200, resp.content)
        entries = resp.json()["entries"]
        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0]["month"], "2026-01-01")  # snapped to 1st of month
        # Replacing again wipes the old rows.
        resp = self.client.put(f"/api/projects/{self.project.id}/cashflow/",
                               [{"month": "2026-03-01", "planned": 50, "actual": 0}],
                               content_type="application/json")
        self.assertEqual(len(resp.json()["entries"]), 1)

    # ── Cash-flow import ──────────────────────────────────────────────────
    def _xlsx_upload(self, build, name="cashflow.xlsx"):
        """Build an in-memory workbook via `build(ws)` and wrap it for upload."""
        import io
        import openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        build(wb.active)
        buf = io.BytesIO()
        wb.save(buf)
        return SimpleUploadedFile(
            name, buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_cashflow_import_wide_layout(self):
        import datetime
        months = [datetime.date(2026, 1, 1), datetime.date(2026, 2, 1), datetime.date(2026, 3, 1)]

        def build(ws):
            ws["A1"] = "Description"
            for i, m in enumerate(months):
                ws.cell(row=1, column=2 + i, value=datetime.datetime(m.year, m.month, 1))
            ws["A2"] = "Planned Cash In /Month"
            for i, v in enumerate([100, 200, 300]):
                ws.cell(row=2, column=2 + i, value=v)
            ws["A3"] = "Planned Monthly %"          # distractor: must be ignored (%)
            ws["A4"] = "Cumulative Planned Cash In"  # distractor: must be ignored
            ws["A5"] = "Actual cash in/Month"
            for i, v in enumerate([90, 150, 0]):
                ws.cell(row=5, column=2 + i, value=v)

        self.login("fa@acme.com")
        resp = self.client.post(f"/api/projects/{self.project.id}/cashflow/import/",
                                {"file": self._xlsx_upload(build)})
        self.assertEqual(resp.status_code, 200, resp.content)
        self.assertEqual(resp.json()["months"], 3)

        entries = self.client.get(f"/api/projects/{self.project.id}/cashflow/").json()["entries"]
        self.assertEqual([e["month"] for e in entries],
                         ["2026-01-01", "2026-02-01", "2026-03-01"])
        self.assertEqual(entries[0]["planned"], "100.00")
        self.assertEqual(entries[0]["actual"], "90.00")
        self.assertEqual(entries[2]["planned"], "300.00")

    def test_cashflow_import_tall_layout_replaces_existing(self):
        # Seed an existing row that the import must wipe.
        self.login("fa@acme.com")
        self.client.put(f"/api/projects/{self.project.id}/cashflow/",
                        [{"month": "2020-01-01", "planned": 5, "actual": 5}],
                        content_type="application/json")

        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "Month", "Planned", "Actual"
            ws["A2"], ws["B2"], ws["C2"] = "2026-06-01", 400, 350
            ws["A3"], ws["B3"], ws["C3"] = "2026-07-01", 500, 450

        resp = self.client.post(f"/api/projects/{self.project.id}/cashflow/import/",
                                {"file": self._xlsx_upload(build)})
        self.assertEqual(resp.status_code, 200, resp.content)
        entries = self.client.get(f"/api/projects/{self.project.id}/cashflow/").json()["entries"]
        self.assertEqual(len(entries), 2)  # old 2020 row gone
        self.assertEqual(entries[0]["month"], "2026-06-01")

    def test_cashflow_import_requires_manage_finances(self):
        def build(ws):
            ws["A1"], ws["B1"], ws["C1"] = "Month", "Planned", "Actual"
            ws["A2"], ws["B2"], ws["C2"] = "2026-06-01", 1, 1
        self.login("fv@acme.com")  # VIEW_FINANCES only
        resp = self.client.post(f"/api/projects/{self.project.id}/cashflow/import/",
                                {"file": self._xlsx_upload(build)})
        self.assertEqual(resp.status_code, 403)

    def test_cashflow_import_unreadable_layout(self):
        def build(ws):
            ws["A1"], ws["B1"] = "Some", "Sheet"
            ws["A2"], ws["B2"] = "with no", "cashflow"
        self.login("fa@acme.com")
        resp = self.client.post(f"/api/projects/{self.project.id}/cashflow/import/",
                                {"file": self._xlsx_upload(build)})
        self.assertEqual(resp.status_code, 400)

    def test_invoice_crud(self):
        self.login("fa@acme.com")
        resp = self.client.post(f"/api/projects/{self.project.id}/invoices/",
                                {"name": "Extract #1", "value": "1500.50", "date": "2026-02-01"})
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertFalse(resp.json()["has_image"])
        iid = resp.json()["id"]
        self.assertEqual(len(self.client.get(f"/api/projects/{self.project.id}/invoices/").json()), 1)
        self.assertEqual(self.client.delete(f"/api/projects/{self.project.id}/invoices/{iid}/").status_code, 204)

    def test_submittals_gated_by_company_perm(self):
        # A user with the submittals company permission can view AND manage.
        self.login("pl@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{self.project.id}/submittals/").status_code, 200)
        resp = self.client.post(f"/api/projects/{self.project.id}/submittals/",
                                {"title": "Door schedule", "submittal_type": "shop_drawing",
                                 "discipline": "architecture", "status": "approved"})
        self.assertEqual(resp.status_code, 201, resp.content)
        self.assertEqual(resp.json()["status_display"], "Approved")
        # A member without the submittals perm is denied.
        self.login("fv@acme.com")
        self.assertEqual(self.client.get(f"/api/projects/{self.project.id}/submittals/").status_code, 403)
