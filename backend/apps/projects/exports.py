"""Primavera P6 export.

Preferred path: return the project's ORIGINAL imported workbook unchanged except
for its progress column(s), refreshed to Planex's current accepted progress.
Every other cell, sheet, formula, and macro is left byte-identical — so the
export matches the reference exactly. Two source shapes are handled, tried in
this order (exact match first, whenever it's usable at all — see
refresh_source_workbook for why the order matters):

* The real P6 schedule template (p6_schedule_import.py / p6_id_schedule_import.py)
  — rows are matched by exact Activity ID, which we kept as Activity.code at
  import time, so refresh just updates each leaf row's "Activity % Complete"
  cell. Tried first: unambiguous whenever the sheet has this header shape,
  regardless of what the sheet itself is named.
* The legacy 'FOR (P6)' outline sheet (p6_import.py, no Start/Finish columns
  so the above never matches it) — rows are matched by (building code +
  normalised task name) instead, a fuzzier fallback for a shape that doesn't
  give us anything more precise.

Either way, unmatched rows keep their original value.

Fallback (no stored workbook): a generated activity sheet in the same column
shape — see build_p6_workbook.
"""
import re
import unicodedata
from collections import defaultdict
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- matching helpers -----------------------------------------------------

_BUILDING_RX = re.compile(r"([A-Za-z]{1,3}\d{1,4})")
_ACTIVITY_ID_RX = re.compile(r"^[A-Za-z]{2,5}\d?-")
_TASHKEEL = dict.fromkeys(range(0x064B, 0x0653))


def _norm(text) -> str:
    """Normalise an Arabic/Latin label for fuzzy equality: drop diacritics,
    unify alef/ya/ta-marbuta forms, collapse whitespace, casefold."""
    s = unicodedata.normalize("NFKC", str(text or "")).translate(_TASHKEEL)
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ى", "ي").replace("ة", "ه")
    s = re.sub(r"\s+", " ", s).strip().casefold()
    return s


def _building_code(text) -> str:
    m = _BUILDING_RX.search(str(text or ""))
    return m.group(1).upper() if m else ""


def _progress_lookup(project) -> dict:
    """(building_code, normalised name) -> list of accepted progress values (0–100),
    one per Planex activity, in import order.

    A task that repeats across a building (e.g. several identical 'وجه 2' rows)
    has several P6 rows AND several Planex activities; we map them positionally so
    each P6 row gets its own activity's value, instead of collapsing them to one
    (averaging) and writing the same number onto every row."""
    scopes = {sid: (name, pid) for sid, name, pid
              in project.scopes.values_list("id", "name", "parent_id")}

    def building_for(scope_id):
        node = scope_id
        while node is not None:
            name, pid = scopes.get(node, ("", None))
            code = _building_code(name)
            if code:
                return code
            node = pid
        return ""

    lookup = defaultdict(list)
    # created_at, id ≈ the order rows were read from the source grid, which lines
    # up with the order the same rows appear in the P6 sheet.
    for sid, name, p in (project.activities.order_by("created_at", "id")
                         .values_list("scope_id", "name", "progress_percent")):
        lookup[(building_for(sid), _norm(name))].append(float(p))
    return lookup


# --- refresh the original workbook ----------------------------------------

# Reference layout: A=Activity ID, B=Activity Name, C=Activity Complete %.
_ID_COL, _NAME_COL, _PCT_COL = 1, 2, 3
P6_SHEET_NAMES = {"for (p6)", "for(p6)", "p6"}


def _find_p6_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower() in P6_SHEET_NAMES:
            return wb[name]
    return None


def _refresh_p6_schedule(wb, project) -> bool:
    """If any sheet in `wb` matches the real P6 schedule template, refresh each
    leaf row's "Activity % Complete" cell from the matching Activity (looked up
    by its exact original Activity ID, kept as Activity.code at import time).
    Returns whether a matching sheet was found (and refreshed)."""
    from .p6_schedule_import import _locate_header

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        located = _locate_header(rows)
        if not located:
            continue
        header_idx, cols = located
        id_col, name_col = cols["activity id"] + 1, cols["activity name"] + 1
        pct_col = cols["activity % complete"] + 1

        progress_by_code = dict(project.activities.exclude(code="").values_list("code", "progress_percent"))
        for r in range(header_idx + 2, ws.max_row + 1):
            a = ws.cell(row=r, column=id_col).value
            b = ws.cell(row=r, column=name_col).value
            if a is None or not isinstance(b, str) or not b.strip():
                continue  # a WBS banner row, not a leaf activity
            pct = progress_by_code.get(str(a).strip()[:60])
            if pct is None:
                continue
            ws.cell(row=r, column=pct_col).value = round(float(pct) / 100, 4)
        return True
    return False


def refresh_source_workbook(project) -> tuple[bytes, str] | None:
    """Reload the stored workbook, refresh the P6 '% Complete' column(s) from
    live progress, and return (bytes, filename). None if there's no stored
    workbook or it matches neither known P6 shape."""
    field = project.source_workbook
    if not field:
        return None
    try:
        raw = field.read()
    except Exception:
        return None

    is_xlsm = field.name.lower().endswith(".xlsm")
    ext = "xlsm" if is_xlsm else "xlsx"
    wb = openpyxl.load_workbook(BytesIO(raw), keep_vba=is_xlsm)

    # Exact-Activity-ID match first: it's unambiguous whenever the sheet has
    # the real P6 header shape (Activity ID/Name/Start/Finish/% Complete),
    # true for anything p6_schedule_import.py/p6_id_schedule_import.py can
    # parse — which includes files whose sheet happens to be literally named
    # "p6", so this must run before the name-based fuzzy path below, or it
    # never gets a chance to. Confirmed: a real Planex-Code-scheme export's
    # sheet is named "p6", which used to short-circuit straight into the
    # fuzzy matcher below — and that matcher's "building code" comes from a
    # regex over raw text (WBS banner rows / literal scope names) tuned for
    # the OLD raw-Arabic-heading shape ("A6", "A15", ...); against the new
    # scheme's own scope names ("PH1", "Z(A)", "Building 6") it either
    # collapses unrelated buildings into one bucket or matches nothing at
    # all, silently leaving every row's % unchanged either way.
    if _refresh_p6_schedule(wb, project):
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"{project.name} - P6.{ext}"

    ws = _find_p6_sheet(wb)
    if ws is not None:
        lookup = _progress_lookup(project)
        seen = defaultdict(int)  # how many P6 rows of each (building, name) we've consumed
        current_building = ""
        for row in ws.iter_rows():
            a = row[_ID_COL - 1].value if len(row) >= _ID_COL else None
            b = row[_NAME_COL - 1].value if len(row) >= _NAME_COL else None
            a_str = str(a).strip() if a is not None else ""

            is_activity = bool(_ACTIVITY_ID_RX.match(a_str)) and b is not None
            if not is_activity:
                # WBS banner row — update the building context when it names one.
                code = _building_code(a_str)
                if code:
                    current_building = code
                continue

            key = (current_building, _norm(b))
            values = lookup.get(key)
            if values and len(row) >= _PCT_COL:
                i = seen[key]
                seen[key] += 1
                # Nth P6 row of this task -> Nth Planex activity (clamp if counts differ).
                value = values[i] if i < len(values) else values[-1]
                # Store as a 0–1 fraction (the sheet's convention); keep the cell's format.
                row[_PCT_COL - 1].value = round(value / 100, 4)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue(), f"{project.name} - FOR (P6).{ext}"

    return None


# --- fallback generated sheet (no stored workbook) ------------------------

HEADERS = ["Activity ID", "Activity Name", "Activity Complete %"]
_WIDTHS = [22, 60, 18]


def _children_map(project):
    by_parent = defaultdict(list)
    for s in project.scopes.all().order_by("sort_order", "name"):
        by_parent[s.parent_id].append(s)
    return by_parent


def _activities_by_scope(project):
    acts = defaultdict(list)
    for a in project.activities.all().order_by("sort_order", "name"):
        acts[a.scope_id].append(a)
    return acts


def build_p6_workbook(project) -> bytes:
    """Generated P6-shaped sheet (Activity ID / Name / Complete %) with the WBS
    as indented group rows. Used when no original workbook was stored."""
    by_parent = _children_map(project)
    acts = _activities_by_scope(project)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FOR (P6)"

    ws.cell(row=1, column=1, value=project.name).font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="5B4FE9")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS[col - 1]
    ws.freeze_panes = "A3"

    group_fill = PatternFill("solid", fgColor="EFEDFE")
    state = {"row": 3}

    def emit_scope(scope, depth):
        r = state["row"]
        cell = ws.cell(row=r, column=1, value=scope.name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(indent=depth)
        for c in (1, 2, 3):
            ws.cell(row=r, column=c).fill = group_fill
        state["row"] += 1
        for child in by_parent.get(scope.id, []):
            emit_scope(child, depth + 1)
        for a in acts.get(scope.id, []):
            rr = state["row"]
            ws.cell(row=rr, column=1, value=a.code or str(a.id)[:8])
            ws.cell(row=rr, column=2, value=a.name).alignment = Alignment(indent=depth + 1)
            pct = ws.cell(row=rr, column=3, value=round(float(a.progress_percent) / 100, 4))
            pct.number_format = "0%"
            state["row"] += 1

    for top in by_parent.get(None, []):
        emit_scope(top, 0)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
