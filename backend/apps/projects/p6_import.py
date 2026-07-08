"""Import a Primavera 'FOR (P6)' sheet into the scope tree.

Fallback for when the workbook has no zone-matrix sheets. The P6 sheet is a flat
WBS export where Excel ROW OUTLINE LEVELS encode the tree depth:
  • group row  — a name in col A, no Activity Name / %  → a WBS node (scope)
  • activity   — Activity ID in col A, name in col B, Complete % (0–1) in col C
There is no weight column, so every activity imports with weight 1 (equal weight).

We anchor on the "Zone …" WBS nodes — the same thing the zone sheets are named
after. The node that directly contains zones is the Stage (a P6 project stage,
e.g. "المرحلة الاولي"); Stages become the top level, so zones that repeat across
stages stay distinguishable. Any title / execution wrappers above the stages are
dropped. Below a zone: a node that holds activities is a Phase, anything between
the zone and a phase is an Area. If there's no stage grouping, zones become the
top level instead.
"""
from collections import defaultdict

import openpyxl

from .models import Activity, ProjectScope

P6_SHEET_NAMES = {"for (p6)"}


def _find_p6_ws(wb):
    for ws in wb.worksheets:
        if ws.title.strip().lower() in P6_SHEET_NAMES:
            return ws
    return None


def parse_p6_tree(file_obj):
    """Return root scope dicts [{name, children[], activities[]}] from the FOR (P6)
    sheet, or None if there's no such sheet / no activities in it."""
    from .imports import _to_pct

    wb = openpyxl.load_workbook(file_obj, data_only=True)  # not read_only: need outline levels
    try:
        ws = _find_p6_ws(wb)
        if ws is None:
            return None
        roots, stack = [], []  # stack of (outline_level, node)
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if a is None or (isinstance(a, str) and not a.strip()):
                continue
            a_str = str(a).strip()
            if a_str.lower() == "activity id":
                continue  # header row
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            rd = ws.row_dimensions.get(r)
            lvl = (rd.outlineLevel if rd is not None else 0) or 0

            if isinstance(b, str) and b.strip():  # a leaf activity (has a name)
                if not stack:
                    continue  # activity with no parent group — skip
                stack[-1][1]["activities"].append({
                    "code": a_str[:60], "name": b.strip()[:200],
                    "pct": _to_pct(c) if isinstance(c, (int, float)) and not isinstance(c, bool) else 0.0,
                })
            else:  # a WBS group node
                node = {"name": a_str[:180], "children": [], "activities": []}
                while stack and stack[-1][0] >= lvl:
                    stack.pop()
                (stack[-1][1]["children"] if stack else roots).append(node)
                stack.append((lvl, node))
    finally:
        wb.close()

    _prune_empty(roots)
    return roots or None


def _has_activities(node):
    """Keep only branches that ultimately hold activities (drops empty WBS like
    a milestones header). Mutates children in place."""
    node["children"] = [c for c in node["children"] if _has_activities(c)]
    return bool(node["activities"] or node["children"])


def _prune_empty(roots):
    roots[:] = [r for r in roots if _has_activities(r)]


def _is_zone(node):
    return "zone" in node["name"].lower()


def _entry_nodes(roots):
    """Where to root the imported tree, and whether those roots are Stages.

    A Stage is a node whose direct children are zones. Stages become the top
    level (so a zone repeated across stages stays distinct). If there are no
    stages, we root at the zone nodes themselves; if there are no zones either,
    we root at the given roots."""
    stages, zones = [], []

    def walk(n):
        if any(_is_zone(c) for c in n["children"]):
            stages.append(n)
            return  # don't descend past a stage
        if _is_zone(n):
            zones.append(n)
            return  # don't descend past a zone
        for c in n["children"]:
            walk(c)

    for r in roots:
        walk(r)
    if stages:
        return stages, True
    return (zones or roots), False


def build_from_p6(project, roots, *, replace=True, snapshot_date=None, source=""):
    """Create scopes + activities from a parsed P6 tree and snapshot progress.
    Stages (or zones, if there are no stages) become the top-level scopes; below
    a zone a node holding activities is a Phase and anything in between is an Area."""
    from django.utils import timezone

    from .imports import _guess_discipline, _save_snapshot, parse_date_from_name
    from .services import project_overall_progress

    Scope = ProjectScope
    company = project.company
    if replace:
        project.scopes.all().delete()

    entries, roots_are_stages = _entry_nodes(roots)
    scopes_by_depth = defaultdict(list)
    activities = []
    counts = defaultdict(int)
    row_counter = [0]

    def type_of(node, forced):
        if forced:
            return forced
        if _is_zone(node):
            return Scope.ScopeType.ZONE
        if node["activities"]:
            return Scope.ScopeType.PHASE  # holds tasks → a work package / phase
        return Scope.ScopeType.AREA       # a grouping between zone and phase

    def walk(node, parent, depth, forced):
        stype = type_of(node, forced)
        counts[stype] += 1
        scope = Scope(company=company, project=project, parent=parent, scope_type=stype,
                      name=node["name"], sort_order=len(scopes_by_depth[depth]),
                      discipline=_guess_discipline(node["name"]) if stype == Scope.ScopeType.PHASE else "")
        scopes_by_depth[depth].append(scope)
        for task in node["activities"]:
            row_counter[0] += 1
            activities.append(Activity(
                company=company, project=project, scope=scope,
                name=task["name"], code=task["code"], weight=1,
                progress_percent=task["pct"], phase_name=node["name"],
                row_index=row_counter[0], progress_type=Activity.ProgressType.PERCENTAGE,
            ))
        for child in node["children"]:
            walk(child, scope, depth + 1, None)

    root_type = Scope.ScopeType.STAGE if roots_are_stages else Scope.ScopeType.ZONE
    for entry in entries:
        walk(entry, None, 0, root_type)

    # Parents before children (UUID PKs are generated in Python, so we only need
    # insert order to satisfy the FK).
    for depth in sorted(scopes_by_depth):
        Scope.objects.bulk_create(scopes_by_depth[depth], batch_size=1000)
    Activity.objects.bulk_create(activities, batch_size=2000)

    snap_date = snapshot_date or parse_date_from_name(source) or timezone.now().date()
    _save_snapshot(project, date=snap_date, source=source)

    return {
        "stages": counts.get(Scope.ScopeType.STAGE, 0),
        "zones": counts.get(Scope.ScopeType.ZONE, 0),
        "subzones": counts.get(Scope.ScopeType.AREA, 0),
        "phases": counts.get(Scope.ScopeType.PHASE, 0),
        "activities": len(activities),
        "overall_progress": project_overall_progress(project),
        "snapshot_date": snap_date.isoformat(),
        "source_kind": "p6",
    }
