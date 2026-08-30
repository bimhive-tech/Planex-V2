"""Excel import for zone-based progress trackers.

Each ZONE sheet is a matrix: subzones across the columns, tasks down the rows,
and a progress cell at every (task, subzone). We import it as:
    Zone (scope)  ->  Subzone (area scope, one per column)  ->  Activity (one per
    (task, subzone) cell, grouped by row_index into task rows).

Layout is detected per sheet (the subzone-label row, the name column, and an
optional leading weight column). Phase/summary rows (col-A "W") are skipped.
The Primavera 'FOR (P6)' and 'Summary' sheets are skipped in this version.
"""
import datetime
import re

import openpyxl
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone

from .models import Activity, ProgressSnapshot, ProjectScope
from .services import project_overall_progress, scope_progress_map

# Dates embedded in tracker file names, e.g. "... - 3-Mar-2026.xlsm" or "2026-03-03".
# The month group captures the first 3 letters (Mar/March both -> "Mar" -> %b).
_DATE_PATTERNS = [
    (re.compile(r"(\d{1,2})[-_ ]([A-Za-z]{3})[a-z]*[-_ ](\d{4})"), "%d %b %Y"),
    (re.compile(r"(\d{4})[-_](\d{1,2})[-_](\d{1,2})"), "%Y %m %d"),
]


def parse_date_from_name(name: str):
    for rx, fmt in _DATE_PATTERNS:
        m = rx.search(name or "")
        if m:
            try:
                return datetime.datetime.strptime(" ".join(m.groups()), fmt).date()
            except ValueError:
                continue
    return None

SKIP_SHEETS = {"for (p6)", "summary"}
MAX_TASKS_PER_ZONE = 2000
MAX_SUBZONES_PER_ZONE = 300


# Phase names in these trackers are usually already one trade's work package
# (e.g. "الاعمال الكهربائية") — a quick keyword guess saves having to tag
# hundreds of imported phases by hand. Blank ("") means unclassified; the
# user can still correct it via the phase's edit form.
_DISCIPLINE_KEYWORDS = {
    ProjectScope.Discipline.CONCRETE: ["خرسان", "حفر", "اساسات", "هيكل", "concrete", "structure"],
    ProjectScope.Discipline.ARCHITECTURE: [
        "تشطيب", "بياض", "دهان", "سيراميك", "رخام", "نجارة", "حدادة", "بلاط", "ارضيات",
        "architecture", "finish", "stair", "entrance",
    ],
    ProjectScope.Discipline.ELECTRICAL: [
        "كهرب", "تيار خفيف", "اضاءة", "انارة", "electrical", "elec",
    ],
    ProjectScope.Discipline.MECHANICAL: [
        "صحي", "صرف", "تكييف", "ميكانيك", "حريق", "مكافحة الحريق", "تهوية", "مياه", "ري",
        "mechanical", "plumbing", "hvac", "elevator", "f.fighting", "fire fighting",
    ],
}


def _guess_discipline(phase_name: str) -> str:
    name = (phase_name or "").lower()
    for discipline, keywords in _DISCIPLINE_KEYWORDS.items():
        if any(k in name for k in keywords):
            return discipline
    return ""


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _to_pct(v):
    f = float(v)
    pct = f * 100 if f <= 1.0001 else f  # values are fractions (0–1)
    return max(0.0, min(100.0, round(pct, 2)))


def _contiguous_runs(row):
    runs, start = [], None
    for i, v in enumerate(row):
        empty = v is None or v == ""
        if not empty and start is None:
            start = i
        elif empty and start is not None:
            runs.append((start, i - 1))
            start = None
    if start is not None:
        runs.append((start, len(row) - 1))
    return runs


def _detect_label_row(rows):
    """Find the subzone-label row: the row whose longest non-empty run has the
    most *string* cells (the index row above it is numeric, so it loses)."""
    best = None  # (string_count, start, end, row_index)
    for i, row in enumerate(rows[:8]):
        for a, b in _contiguous_runs(row):
            strings = sum(1 for c in range(a, b + 1) if isinstance(row[c], str) and row[c].strip())
            if strings >= 2 and (best is None or strings > best[0]):
                best = (strings, a, b, i)
    return best


def parse_sheet(rows, is_header=None):
    """Return {subzones: [labels], tasks: [{name, weight, phase, row_index, cells}]}
    where cells is a list aligned to subzones (None for blanks).

    `is_header(row_idx, name_col)` (optional) flags a styled section/phase header
    row — the trackers only mark the FIRST discipline with a col-A "W", styling
    the rest (bold + fill) instead, so without this every later discipline's
    tasks would collapse into the first phase."""
    det = _detect_label_row(rows)
    if not det:
        return None
    _, sub_start, sub_end, label_row = det
    subzone_cols = list(range(sub_start, min(sub_end + 1, sub_start + MAX_SUBZONES_PER_ZONE)))
    name_col = sub_start - 1
    if name_col < 0:
        return None
    weight_col = name_col - 1 if name_col - 1 >= 0 else None

    subzones = [str(rows[label_row][c]).strip() for c in subzone_cols]

    tasks, phase, ri = [], "", 0
    skip_first_summary = weight_col is None  # no "W" marker -> first row is the summary
    for offset, row in enumerate(rows[label_row + 1:]):
        idx = label_row + 1 + offset  # absolute 0-based row index (for style lookups)
        if len(tasks) >= MAX_TASKS_PER_ZONE:
            break
        wcell = row[weight_col] if (weight_col is not None and weight_col < len(row)) else None
        name = row[name_col] if name_col < len(row) else None

        # Phase / section header row — col-A "W", or a styled (bold + filled) name
        # cell for the disciplines that carry no "W".
        is_phase = isinstance(wcell, str) and wcell.strip().upper() == "W"
        if not is_phase and is_header is not None and isinstance(name, str) and name.strip():
            is_phase = is_header(idx, name_col)
        if is_phase:
            if isinstance(name, str) and name.strip():
                phase = name.strip()[:180]
            continue
        if not isinstance(name, str) or not name.strip():
            continue
        if skip_first_summary:
            # No "W" column on this sheet — the first named row is the phase/summary.
            skip_first_summary = False
            phase = name.strip()[:180]
            continue

        cells, any_num = [], False
        for c in subzone_cols:
            v = row[c] if c < len(row) else None
            if _is_num(v):
                cells.append(_to_pct(v))
                any_num = True
            else:
                cells.append(None)
        if not any_num:
            continue

        weight = float(wcell) if (_is_num(wcell) and wcell > 0) else 1.0
        ri += 1
        tasks.append({"name": name.strip()[:200], "weight": weight, "phase": phase,
                      "row_index": ri, "cells": cells})
    return {"subzones": subzones, "tasks": tasks}


def _styled_header(cell):
    """True when a cell is a styled section header (bold + a solid fill). Works on
    read-only cells too, so we can detect phase headers without leaving streaming
    mode — a full-workbook load of a 20MB+ tracker needs ~1GB and OOMs the worker."""
    try:
        return bool(cell.font and cell.font.bold) and bool(cell.fill and cell.fill.patternType)
    except (AttributeError, ValueError):
        return False


def parse_workbook(file_obj) -> dict:
    """Convenience wrapper: open `file_obj` read-only and parse its zone sheets.
    Callers already holding an open workbook should use parse_workbook_sheets —
    opening a 20MB+ tracker costs ~26s, so the import path shares one open."""
    # read_only streams the file (keeps memory sane on big workbooks). ReadOnlyCell
    # still exposes font/fill, so styled phase headers are detectable here — and we
    # only iterate the zone sheets, never the huge skipped FOR (P6)/Summary sheets.
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    try:
        return parse_workbook_sheets(wb)
    finally:
        wb.close()


def parse_workbook_sheets(wb) -> dict:
    """Parse every zone-matrix sheet in an already-open read-only workbook."""
    result = {}
    for name in wb.sheetnames:
        if name.strip().lower() in SKIP_SHEETS:
            continue
        ws = wb[name]
        values, styled = [], []
        for row in ws.iter_rows():
            values.append(tuple(c.value for c in row))
            styled.append(tuple(_styled_header(c) for c in row))

        def is_header(row_idx0, name_col0, styled=styled):
            return name_col0 < len(styled[row_idx0]) and styled[row_idx0][name_col0]

        sheet = parse_sheet(values, is_header)
        if sheet and sheet["tasks"] and sheet["subzones"]:
            result[name.strip()] = sheet
    return result


def _save_snapshot(project, *, date, source, schedule_import=None):
    """Capture the project's aggregate progress as a dated snapshot (upsert by
    date). `schedule_import` scopes every aggregate to one batch — normally
    the batch this same import just created, so a re-import's snapshot
    reflects only its own fresh activities, not every batch ever imported
    combined (batches are no longer deleted on re-import — see
    ScheduleImport's own docstring). Falls back to "whatever's current" when
    not given, matching every pre-existing caller."""
    from .services import latest_schedule_import

    if schedule_import is None:
        schedule_import = latest_schedule_import(project)
    activities = project.activities.filter(schedule_import=schedule_import) if schedule_import else project.activities.all()
    scopes = project.scopes.filter(schedule_import=schedule_import) if schedule_import else project.scopes.all()
    agg = activities.aggregate(
        total=Count("id"),
        completed=Count("id", filter=Q(progress_percent__gte=100)),
        not_started=Count("id", filter=Q(progress_percent__lte=0)),
    )
    total = agg["total"]
    breakdown = {
        "total": total, "completed": agg["completed"], "not_started": agg["not_started"],
        "in_progress": total - agg["completed"] - agg["not_started"],
    }
    progress = scope_progress_map(project, schedule_import=schedule_import)
    zones = [
        {"name": z.name, "progress": progress.get(str(z.id), 0.0)}
        for z in scopes.filter(scope_type=ProjectScope.ScopeType.ZONE).order_by("sort_order")
    ]
    ProgressSnapshot.objects.update_or_create(
        project=project, date=date,
        defaults={"company": project.company,
                  "overall_progress": project_overall_progress(project, schedule_import=schedule_import),
                  "breakdown": breakdown, "zones": zones, "scopes": progress, "source": source[:200]},
    )


@transaction.atomic
def import_workbook(project, file_obj, *, replace=True, snapshot_date=None, source="") -> dict:
    # A prior P6 import may have left authoritative overall-% figures on the
    # project (see project_overall_progress / reports._planned_progress). Clear
    # them up front, before format dispatch: if this import turns out to be
    # another P6 schedule, that path sets fresh figures below; if it's a zone
    # tracker or the legacy P6 fallback instead, neither knows about these
    # fields, and leaving the old project's numbers in place would silently
    # outlive the tree they described.
    if replace and (project.imported_progress_percent is not None
                    or project.imported_planned_progress_percent is not None):
        project.imported_progress_percent = None
        project.imported_planned_progress_percent = None
        project.save(update_fields=[
            "imported_progress_percent", "imported_planned_progress_percent", "updated_at",
        ])

    # Both read-only probes share ONE open: on a 20MB+ tracker openpyxl spends
    # ~26s just opening the file, so probing formats with an open apiece is the
    # dominant cost of an import.
    from .p6_id_schedule_import import parse_id_schedule_sheets
    from .p6_schedule_import import build_from_p6_schedule, parse_p6_schedule_sheets
    try:
        file_obj.seek(0)
    except (AttributeError, OSError):
        pass
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    try:
        # Newer exports carry the WBS path as a segmented Activity ID (see
        # p6_id_schedule_import) instead of leading-space indentation — tried
        # first since its detection is stricter (samples the ID column's own
        # values) and would never accidentally match a leading-space file.
        schedule_roots = parse_id_schedule_sheets(wb)
        is_segmented_id = bool(schedule_roots)
        if not schedule_roots:
            # The real P6 schedule export (Activity ID/Name/Start/Finish/%
            # Complete, WBS via indentation) is the standard template going
            # forward — probed next since it never looks like a zone matrix
            # and would otherwise misdetect as one.
            schedule_roots = parse_p6_schedule_sheets(wb)
        parsed = {} if schedule_roots else parse_workbook_sheets(wb)
    finally:
        wb.close()

    if schedule_roots:
        return build_from_p6_schedule(project, schedule_roots,
                                      snapshot_date=snapshot_date, source=source,
                                      unwrap_single_root=not is_segmented_id)

    if not parsed:
        # No zone-matrix sheets — fall back to the legacy 'FOR (P6)' sheet if the
        # workbook has one, so an older P6-only export still imports.
        from .p6_import import build_from_p6, parse_p6_tree
        try:
            file_obj.seek(0)
        except (AttributeError, OSError):
            pass
        roots = parse_p6_tree(file_obj)
        if roots:
            return build_from_p6(project, roots, snapshot_date=snapshot_date, source=source)
        return {"zones": 0, "subzones": 0, "activities": 0, "overall_progress": 0.0,
                "error": "No zone sheets or FOR (P6) sheet recognised."}

    # A re-import creates a new, permanently-retained batch instead of
    # deleting the previous one — see ScheduleImport's own docstring. `date`
    # is resolved up front (not after building) since every new row is
    # tagged to this batch as it's constructed.
    from .models import ScheduleImport

    snap_date = snapshot_date or parse_date_from_name(source) or timezone.now().date()
    company = project.company
    schedule_import = ScheduleImport.objects.create(company=company, project=project, date=snap_date, source=source)

    Scope = ProjectScope
    zones, subz, phases, activities = [], [], [], []
    subzone_total = 0
    for z, (zone_name, sheet) in enumerate(parsed.items()):
        zone = Scope(company=company, project=project, schedule_import=schedule_import,
                     scope_type=Scope.ScopeType.ZONE, name=zone_name, sort_order=z)
        zones.append(zone)

        # Group tasks by phase (preserving order). Tree is
        # Zone -> Subzone -> Phase -> Task: each subzone holds the phases, each phase
        # holds that subzone's task cells (an Activity per cell).
        order, by_phase = [], {}
        for task in sheet["tasks"]:
            ph = task["phase"] or "Tasks"
            if ph not in by_phase:
                by_phase[ph] = []
                order.append(ph)
            by_phase[ph].append(task)

        for c, label in enumerate(sheet["subzones"]):
            subzone = Scope(company=company, project=project, parent=zone, schedule_import=schedule_import,
                            scope_type=Scope.ScopeType.AREA, name=label or f"SZ{c + 1}", sort_order=c)
            subz.append(subzone)
            subzone_total += 1
            for pi, ph in enumerate(order):
                phase = Scope(company=company, project=project, parent=subzone, schedule_import=schedule_import,
                              scope_type=Scope.ScopeType.PHASE, name=ph, sort_order=pi,
                              discipline=_guess_discipline(ph))
                phases.append(phase)
                for task in by_phase[ph]:
                    val = task["cells"][c]
                    if val is None:
                        continue
                    activities.append(Activity(
                        company=company, project=project, scope=phase, schedule_import=schedule_import,
                        name=task["name"], weight=task["weight"], progress_percent=val,
                        phase_name=ph, row_index=task["row_index"],
                        subzone_code=label or f"SZ{c + 1}", subzone_index=c,
                        progress_type=Activity.ProgressType.PERCENTAGE,
                    ))

    # Insert parents before children (UUID PKs are generated client-side).
    Scope.objects.bulk_create(zones, batch_size=1000)
    Scope.objects.bulk_create(subz, batch_size=1000)
    Scope.objects.bulk_create(phases, batch_size=1000)
    Activity.objects.bulk_create(activities, batch_size=2000)
    schedule_import.activity_count = len(activities)
    schedule_import.save(update_fields=["activity_count", "updated_at"])

    _save_snapshot(project, date=snap_date, source=source, schedule_import=schedule_import)

    return {
        "zones": len(zones),
        "subzones": subzone_total,
        "phases": len(phases),
        "activities": len(activities),
        "overall_progress": project_overall_progress(project, schedule_import=schedule_import),
        "snapshot_date": snap_date.isoformat(),
        "schedule_import_id": str(schedule_import.id),
    }


