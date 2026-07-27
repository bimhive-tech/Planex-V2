"""Import monthly cash flow, and invoices/extracts, from an Excel workbook.

Cash flow supports two layouts:

* WIDE (the common site-office / Primavera cash-flow sheet): months run across a
  header row as real dates, with labelled "planned" and "actual" cash rows below.
* TALL (a simple template): Month | Planned | Actual columns down the rows.

Cumulative and percentage rows are ignored — we only want the per-month amounts,
since the app stores those and charts the cumulative S-curve itself.

Invoices come from a different shape entirely — see parse_invoice_extracts below.
"""
import datetime
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction

from .models import CashFlowEntry, Invoice

SCAN_ROWS = 100          # how deep to look for the header / label rows
SCAN_COLS = 200          # cap width so 16k-column export sheets don't stall us
MIN_MONTHS = 3           # a valid cash-flow needs at least a few months
_EXCLUDE = ("cumulative", "cumm", "%", "percent")  # skip running-total / % rows


_MONTH_STR_FORMATS = ("%Y-%m-%d", "%Y-%m", "%d/%m/%Y", "%b %Y", "%B %Y")


def _as_month(value, strings=False):
    """Coerce a cell to the first day of its month, or None if it isn't a date.

    `strings=True` also parses common textual dates ("2026-06-01", "Jun 2026") —
    used only by the tall-template reader, so stray text like a manpower "OCT"
    header can't masquerade as a month in the wide reader's detection."""
    if isinstance(value, datetime.datetime):
        return value.date().replace(day=1)
    if isinstance(value, datetime.date):
        return value.replace(day=1)
    if strings and isinstance(value, str):
        text = value.strip()
        for fmt in _MONTH_STR_FORMATS:
            try:
                return datetime.datetime.strptime(text, fmt).date().replace(day=1)
            except ValueError:
                continue
    return None


def _as_amount(value):
    """Coerce a cell to a 2dp Decimal, or None if it isn't a plain number."""
    if value is None or isinstance(value, (datetime.datetime, datetime.date, str)):
        # strings are rejected: a label column must not be read as an amount
        if isinstance(value, str):
            value = value.strip().replace(",", "")
            if not value:
                return None
        else:
            return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _row_label(ws, row):
    """The first text cell in the first few columns — the row's label."""
    for col in range(1, 5):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and v.strip():
            return v.strip().lower()
    return ""


def _is_data_label(label, keyword):
    return keyword in label and not any(x in label for x in _EXCLUDE)


def _read_wide(ws):
    """Find a date-header row and the planned/actual rows beneath it."""
    header_row = header = None
    for row in range(1, min(ws.max_row, SCAN_ROWS) + 1):
        months = {}
        for col in range(1, min(ws.max_column, SCAN_COLS) + 1):
            m = _as_month(ws.cell(row=row, column=col).value)
            if m:
                months[col] = m
        if len(months) >= MIN_MONTHS:
            header_row, header = row, months
            break
    if not header:
        return {}

    planned_row = actual_row = None
    for row in range(header_row, min(ws.max_row, SCAN_ROWS) + 1):
        label = _row_label(ws, row)
        if planned_row is None and _is_data_label(label, "planned"):
            planned_row = row
        # Site cash-flow sheets often call the actual cash-in "invoices" rather
        # than "actual" — treat either as the real-money row.
        elif actual_row is None and (_is_data_label(label, "actual") or _is_data_label(label, "invoice")):
            actual_row = row
    if planned_row is None and actual_row is None:
        return {}

    out = {}
    for col, month in header.items():
        planned = _as_amount(ws.cell(row=planned_row, column=col).value) if planned_row else None
        actual = _as_amount(ws.cell(row=actual_row, column=col).value) if actual_row else None
        if planned is None and actual is None:
            continue  # a header date with no cash under it isn't a real month
        out[month] = (planned or Decimal("0"), actual or Decimal("0"))
    return out


def _read_tall(ws):
    """Find a Month / Planned / Actual column header, then read rows beneath."""
    cols = {}
    header_row = None
    for row in range(1, min(ws.max_row, SCAN_ROWS) + 1):
        found = {}
        for col in range(1, min(ws.max_column, SCAN_COLS) + 1):
            v = ws.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            t = v.strip().lower()
            if t in ("month", "date") and "month" not in found:
                found["month"] = col
            elif _is_data_label(t, "planned") and "planned" not in found:
                found["planned"] = col
            elif _is_data_label(t, "actual") and "actual" not in found:
                found["actual"] = col
        if "month" in found and ("planned" in found or "actual" in found):
            cols, header_row = found, row
            break
    if not header_row:
        return {}

    out = {}
    for row in range(header_row + 1, ws.max_row + 1):
        month = _as_month(ws.cell(row=row, column=cols["month"]).value, strings=True)
        if not month:
            continue
        planned = _as_amount(ws.cell(row=row, column=cols["planned"]).value) if "planned" in cols else None
        actual = _as_amount(ws.cell(row=row, column=cols["actual"]).value) if "actual" in cols else None
        out[month] = (planned or Decimal("0"), actual or Decimal("0"))
    return out


def parse_cashflow(upload):
    """Return {month(date): (planned, actual)} from the best sheet that yields a
    recognisable cash-flow layout, or raise ValueError if none do.

    A workbook with several per-contract cash-flow sheets (e.g. cashflow1/2/3)
    plus one aggregating them (e.g. "cashflow total") should import the total,
    not whichever partial sheet happens to come first — so sheets with "total"
    in their name are tried before the rest."""
    wb = openpyxl.load_workbook(upload, read_only=False, data_only=True)
    try:
        sheets = sorted(wb.worksheets, key=lambda ws: "total" not in ws.title.strip().lower())
        # A reader only returns data once it has locked onto a real layout (the
        # wide reader needs MIN_MONTHS date cells to accept a header row), so any
        # non-empty result here is a genuine cash-flow — even a couple of months.
        for reader in (_read_wide, _read_tall):
            for ws in sheets:
                data = reader(ws)
                if data:
                    return data
    finally:
        wb.close()
    raise ValueError(
        "No cash-flow layout found. Expected either month dates across a row with "
        "'planned'/'actual' rows below, or Month/Planned/Actual columns."
    )


def import_cashflow(project, upload):
    """Replace the project's monthly cash flow from an uploaded workbook.

    Returns a small summary for the UI. Replace (not merge) keeps it predictable:
    what's in the sheet is what you get, matching how the manual grid saves."""
    data = parse_cashflow(upload)
    months = sorted(data)
    rows = [
        CashFlowEntry(company=project.company, project=project, month=m,
                      planned=data[m][0], actual=data[m][1])
        for m in months
    ]
    with transaction.atomic():
        project.cashflow_entries.all().delete()
        CashFlowEntry.objects.bulk_create(rows)
    return {
        "months": len(rows),
        "first_month": months[0].isoformat(),
        "last_month": months[-1].isoformat(),
    }


# --- Invoices / extracts (مستخلصات) ----------------------------------------
#
# The reference layout is NOT a flat invoice list — it's a per-BOQ-item matrix:
# one row per work item/zone, and one column-GROUP per submitted extract
# ("حتى <date>" = "up to <date>"), each group holding a "رقم المستخلص" (extract
# number) column and an "اجمالي الأعمال" (total work value) column. The value in
# that column is the item's CUMULATIVE work done as of that extract, not the
# extract's own amount — it climbs toward the item's contract value.
#
# An invoice's own value is therefore derived, not read directly: sum the
# "اجمالي الأعمال" column down every row to get the project's cumulative total at
# that extract, then subtract the previous extract's cumulative total. That
# matches how a progress "مستخلص" actually works — it bills the work done since
# the last one.

_ARABIC_MONTHS = {
    "يناير": 1, "فبراير": 2, "مارس": 3, "ابريل": 4, "أبريل": 4, "مايو": 5,
    "يونيو": 6, "يوليو": 7, "اغسطس": 8, "أغسطس": 8, "سبتمبر": 9,
    "اكتوبر": 10, "أكتوبر": 10, "نوفمبر": 11, "ديسمبر": 12,
}
_EXTRACT_DATE_RX = re.compile(r"(\d{1,2})\s+([^\s\-]+)\s*-\s*(\d{4})")
_TOTAL_WORKS_LABEL = "اجمالي الأعمال"
_EXTRACT_HEADER_SCAN_ROWS = 10


def _parse_extract_date(label):
    """Best-effort parse of a "حتى 15 ديسمبر - 2023" style label. Real trackers
    have typos (a wrong year on a late column is common) — return None rather
    than raise, so one bad label doesn't block the whole import."""
    m = _EXTRACT_DATE_RX.search(label or "")
    if not m:
        return None
    day, month_name, year = m.groups()
    month = _ARABIC_MONTHS.get(month_name.strip())
    if not month:
        return None
    try:
        return datetime.date(int(year), month, int(day))
    except ValueError:
        return None


def _locate_extract_header(ws):
    """Find (group_row, sub_row): sub_row is the row carrying "اجمالي الأعمال"
    sub-headers; group_row is the nearest row above it carrying the per-extract
    label (the label only occupies the first cell of its merged span — read_only
    cells outside that first cell come back None, same as every other merged
    header in these trackers)."""
    for sub_row in range(1, min(ws.max_row, _EXTRACT_HEADER_SCAN_ROWS) + 1):
        cells = [c.value for c in next(ws.iter_rows(min_row=sub_row, max_row=sub_row))]
        if any(isinstance(v, str) and v.strip() == _TOTAL_WORKS_LABEL for v in cells):
            for group_row in range(sub_row - 1, 0, -1):
                grp = [c.value for c in next(ws.iter_rows(min_row=group_row, max_row=group_row))]
                if any(isinstance(v, str) and v.strip() for v in grp):
                    return group_row, sub_row
    return None


_EXTRACT_NUMBER_LABEL = "رقم المستخلص"
_PLACEHOLDER_VALUES = {"-", "—", ""}


def parse_invoice_extracts(upload):
    """Return ([{name, date, value}], skipped) — one dict per submitted extract
    in chronological order, value already converted from cumulative to the
    extract's own amount; `skipped` counts extracts dropped as unreliable (see
    below). None (not a tuple) if no sheet matches this layout at all."""
    wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            located = _locate_extract_header(ws)
            if not located:
                continue
            group_row, sub_row = located
            group_cells = [c.value for c in next(ws.iter_rows(min_row=group_row, max_row=group_row))]
            sub_cells = [c.value for c in next(ws.iter_rows(min_row=sub_row, max_row=sub_row))]

            # Forward-fill the group label across its merged span, keeping the
            # "اجمالي الأعمال" sub-column of each group — one per extract — and,
            # when present, the "رقم المستخلص" column immediately to its left.
            # That extract-number cell (e.g. "مستخلص جاري (8)") is the real,
            # human-recognizable name of the invoice; the date label above it is
            # just the column heading and reads badly as a name.
            periods, label = [], ""
            for i, v in enumerate(group_cells):
                if isinstance(v, str) and v.strip():
                    label = v.strip()
                if i < len(sub_cells) and isinstance(sub_cells[i], str) \
                        and sub_cells[i].strip() == _TOTAL_WORKS_LABEL and label:
                    number_col = i - 1 if i >= 1 and isinstance(sub_cells[i - 1], str) \
                        and sub_cells[i - 1].strip() == _EXTRACT_NUMBER_LABEL else None
                    periods.append((i, number_col, label))
            if not periods:
                continue

            sums = defaultdict(float)
            numbers = {}  # value_col -> first real "رقم المستخلص" text seen
            for row in ws.iter_rows(min_row=sub_row + 1, values_only=True):
                for idx, number_col, _ in periods:
                    if idx < len(row):
                        v = row[idx]
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            sums[idx] += v
                    if number_col is not None and idx not in numbers and number_col < len(row):
                        n = row[number_col]
                        if isinstance(n, str) and n.strip() and n.strip() not in _PLACEHOLDER_VALUES:
                            numbers[idx] = n.strip()

            # Column order in the sheet is NOT reliable as extract order — the
            # reference file has a late column physically placed after ones
            # dated months later than it (a scope added to the tracker after
            # the fact, appended rather than inserted in date order). Diffing
            # cumulative totals in sheet order on a file like that produces
            # nonsense: a "delta" between two unrelated points in time. Sort by
            # the parsed date first — column position only breaks ties (kept
            # stable) when two extracts share one date.
            dated = [(idx, numbers.get(idx, label), _parse_extract_date(label), sums.get(idx, 0.0))
                    for idx, _, label in periods]
            dated.sort(key=lambda t: (t[2] is None, t[2] or datetime.date.max, t[0]))

            # A cumulative-to-date total must not go backwards. It does, hard, in
            # trackers whose source formulas have quietly broken (this workbook
            # has confirmed #REF! errors elsewhere) — a later column's cached sum
            # can be a stale fraction of an earlier one. Once a column's total
            # comes in below the highest total seen so far, it and everything
            # computed from it is unreliable: drop it rather than book a
            # fabricated invoice, and keep comparing later columns against the
            # last column that *did* make sense.
            seen = defaultdict(int)  # de-dupes a repeated extract name into "(2)", "(3)"...
            out, last_good = [], 0.0
            skipped = 0
            for idx, name, date, cumulative in dated:
                if cumulative < last_good:
                    skipped += 1
                    continue
                seen[name] += 1
                display = name if seen[name] == 1 else f"{name} ({seen[name]})"
                out.append({
                    "name": display[:200], "date": date,
                    "value": round(cumulative - last_good, 2),
                })
                last_good = cumulative
            return out, skipped
    finally:
        wb.close()
    return None


def import_invoices(project, upload):
    """Create/update the project's invoices from an uploaded extract-comparison
    workbook. Upserts by (name, date) rather than replacing wholesale, so a
    re-import (the tracker gets a new extract column each period) never touches
    invoices entered by hand or their attached scan images."""
    result = parse_invoice_extracts(upload)
    if not result:
        raise ValueError(
            "No invoice/extract layout found. Expected a per-item table with "
            "'رقم المستخلص' and 'اجمالي الأعمال' columns, one pair per submitted extract."
        )
    periods, skipped = result
    created = updated = 0
    with transaction.atomic():
        existing = {(inv.name, inv.date): inv for inv in project.invoices.all()}
        for i, p in enumerate(periods):
            value = Decimal(str(p["value"])).quantize(Decimal("0.01"))
            match = existing.get((p["name"], p["date"]))
            if match:
                if match.value != value:
                    match.value = value
                    match.save(update_fields=["value"])
                updated += 1
            else:
                Invoice.objects.create(
                    company=project.company, project=project, name=p["name"],
                    value=value, date=p["date"], sort_order=i,
                )
                created += 1
    return {"periods": len(periods), "created": created, "updated": updated, "skipped": skipped}
